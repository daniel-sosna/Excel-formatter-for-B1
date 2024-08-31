from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime

from utils import col_to_ind
from config import *


class LoadWorkbook():
	def __init__(self, filename, read_only=False):
		try:
			self.workbook = load_workbook(filename, read_only=read_only)
		except Exception as e:
			self.sheet = None
			print(f"🚨 Failed to open \"{filename}\". See the error below:")
			print(type(e), e)
		else:
			print(f"Successfully opened \"{filename}\"")
			self.sheet = self.workbook.active


class WriteSalesToExcel():
	def __init__(self, filename, sales, eu, not_eu):
		self.all = sales
		self.eu = eu
		self.not_eu = not_eu
		self.write(filename)

	def write(self, filename):
		workbook = Workbook()
		workbook.active.title = "Visi"
		self.write_to_sheet(workbook.active, self.all)
		self.write_to_sheet(workbook.create_sheet("ES"), [(k, *v) for k, v in self.eu.items()],
							headers=("Country", "Total without VAT", "VAT", "Total"), sum_start_from_header=2)
		self.write_to_sheet(workbook.create_sheet("ne ES"), self.not_eu)
		try:
			workbook.save(filename)
			print(f"📥︎ Successfully saved sales into \"{filename}\"")
		except Exception as e:
			print(f"🚨 Failed to save sales. See the error below and close the \"{filename}\" file if it is open.")
			print(type(e), e)

	def write_to_sheet(self, sheet, sales, headers=("Date", "Country", "Total"), sum_start_from_header=3):
		sheet.append(headers)
		for cell in sheet[1]:
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal='center')
		for row in sales:
			sheet.append(row)
		self.add_sum_cells_to_sheet(sheet, len(sales), headers, sum_start_from_header)
		self.align_columns_width(sheet, (get_column_letter(len(headers)+2)))

	def add_sum_cells_to_sheet(self, sheet, n_sales, headers, start_header):
		row = 1
		for i, header in enumerate(headers[start_header-1:], start=start_header):
			row += 1
			# Header cell
			h_cell = sheet.cell(row=row, column=len(headers)+2)
			h_cell.value = header + ':'
			h_cell.font = Font(bold=True, color="FF0000")
			h_cell.alignment = Alignment(horizontal='center')
			# Sum cell
			s_cell = sheet.cell(row=row, column=len(headers)+3)
			s_cell.value = f'=SUM({get_column_letter(i)}{2}:{get_column_letter(i)}{n_sales+1})'
			s_cell.font = Font(color="C00000")

	def align_columns_width(self, sheet, columns=('E')):
		max_width = 0
		for col in columns:
			for cell in sheet[col]:
				if cell.value:
					max_width = max(max_width, len(cell.value))
			if max_width > sheet.column_dimensions[col].width:
				sheet.column_dimensions[col].width = max_width


class FillOutTemplateFile():
	def __init__(self, template_filename, result_filename, sales):
		self.sales = sales
		self.fill(template_filename, result_filename)

	def fill(self, template_filename, result_filename):
		wb = LoadWorkbook(template_filename)
		sheet = wb.sheet
		if not sheet:
			return

		for i, (date, country, price) in enumerate(self.sales, start=1):
			sheet.cell(row=i+1, column=col_to_ind(VARIABLES['date'], 1)).value = date
			sheet.cell(row=i+1, column=col_to_ind(VARIABLES['number'], 1)).value = i
			sheet.cell(row=i+1, column=col_to_ind(VARIABLES['country'], 1)).value = country
			sheet.cell(row=i+1, column=col_to_ind(VARIABLES['price'], 1)).value = price

			for col, val in CONSTANTS.items():
				sheet.cell(row=i+1, column=col_to_ind(col, 1)).value = val

		try:
			wb.workbook.save(result_filename)
			print(f"📥︎ Successfully saved sales outside the EU using the template into \"{result_filename}\"")
		except Exception as e:
			print(f"🚨 Failed to save sales outside the EU using the template. See the error below and close the \"{result_filename}\" file if it is open.")
			print(type(e), e)


class SaveData():
	def __init__(self, all_sales, EU_sales, not_EU_sales):
		print("# Saving modified sales data:")
		print("↪ Enter the MONTH you want to appear in the output file names.")
		print(f"Or press Enter to use the default value (which is the current month: {datetime.now().month}).")
		month = input("≫ ")
		print("↪ Enter the YEAR you want to appear in the output file names.")
		print(f"Or press Enter to use the default value (which is the current year: {datetime.now().year}).")
		year = input("≫ ")

		SALES_MONTH = month if month else datetime.now().month
		SALES_YEAR = year if month else datetime.now().year

		WriteSalesToExcel(f'{SALES_OUTPUT}{SALES_YEAR}-{SALES_MONTH}.xlsx', all_sales, EU_sales, not_EU_sales)
		
		print("↪ Enter the path (filename if the file is in the same folder) to the TEMPLATE FILE or drag it into this window.")
		print(f"Or press Enter to use the default value (\"{TEMPLATE_PATH}\").")
		template = input("≫ ")
		template_filename = template if template else TEMPLATE_PATH
		FillOutTemplateFile(template_filename, f'{TEMPLATE_OUTPUT}{SALES_YEAR}-{SALES_MONTH}.xlsx', not_EU_sales)
