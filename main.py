from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment
from vat import EU_VAT

DATE_COL = 'A'
COUNTRY_COL = 'O'
TOTAL_COL = 'X'

def col_to_ind(column:str, start:int=0) -> int:
	''' Converts a column name (e.g. 'A', 'AF', 'CK') to an index '''
	index = start - 1
	for i, letter in enumerate(column[::-1]): # Run through reversed column name
		index += (ord(letter) - ord('A') + 1) * 26**i
	# 'ABC'  ->  'CBA'  ->  'A'*(26^2) + 'B'*(26^1) + 'C'*(26^0)  ->  1*676 + 2*26 + 3*1  ->  731
	return index


class LoadWorkbook():
	def __init__(self, filename):
		try:
			workbook = load_workbook(filename, read_only=True)
		except Exception as e:
			self.sheet = None
			print(f"Failed to open the \"{filename}\" file. See the error below:")
			print(type(e), e)
		else:
			self.open_worksheet(workbook)

	def open_worksheet(self, workbook):
		self.sheet = workbook.active
		print(f"Data from sheet '{self.sheet.title}'")


class DataExtractor():
	def __init__(self, sheet):
		self.sheet = sheet
		self.headers = [self.sheet[col+'1'].value for col in (DATE_COL, COUNTRY_COL, TOTAL_COL)]
		print(f"Columns to parse: {self.headers[0]}, {self.headers[1]}, {self.headers[2]}\n")

	def run(self, start=2, stop=None) -> tuple[list, bool]:
		data = []
		skipped_count = 0
		error_count = 0

		print("# Extracting sales:")
		for i, row in enumerate(self.sheet.iter_rows(min_row=start, values_only=True), start=start):
			if stop and i > stop:
				i -= 1
				break

			row_data, is_valid = self.get_row_data(i, row)

			# Save modified row data if valid
			if is_valid:
				data.append(row_data)
			elif row_data:
				error_count += 1
			else:
				skipped_count += 1

		self.print_results(i - start + 1, len(data), skipped_count, error_count)
		return sorted(data, key=lambda x: x[0]), True if not error_count else False

	def get_row_data(self, i, row) -> tuple[tuple, bool] | tuple[None, False]:
		# Get needed columns data
		date = row[col_to_ind(DATE_COL)]
		country = row[col_to_ind(COUNTRY_COL)]
		total = row[col_to_ind(TOTAL_COL)]

		# Return None if row is blank
		if not (date or country or total):
			print(f"✔ No data in row {i}. Skipped")
			return None, False

		is_row_valid = True

		# [Sale Date]
		if date:
			try:
				(months, day, year_tens) = date.split('/')
				new_date = f'20{year_tens}-{months}-{day}'
			except Exception as e:
				print(f"❌ [{DATE_COL}{i}] Incorrect '{self.headers[0]}' in row {i}: '{date}'")
				new_date = date
				is_row_valid = False
		else:
			print(f"❌ [{DATE_COL}{i}] No '{self.headers[0]}' in row {i}: ({date}, {country}, {total})")
			is_row_valid = False

		# [Ship Country]
		if not country:
			print(f"❌ [{COUNTRY_COL}{i}] No '{self.headers[1]}' in row {i}: ({date}, {country}, {total})")
			is_row_valid = False

		# [Order Total]
		if total:
			if isinstance(total, str):
				try:
					old_total, total = total, float(total.replace(',', ''))
					print(f"✔ [{TOTAL_COL}{i}] Incorrect '{self.headers[2]}' in row {i}: '{old_total}'. Changed to '{total}'")
				except Exception as e:
					print(f"❌ [{TOTAL_COL}{i}] Incorrect '{self.headers[2]}' in row {i}: '{total}'. {type(e)}: {e}")
					is_row_valid = False
		else:
			print(f"❌ [{TOTAL_COL}{i}] No '{self.headers[2]}' in row {i}: ({date}, {country}, {total})")
			is_row_valid = False

		return (
			new_date if 'new_date' in locals() else date,
			country,
			total
		), is_row_valid

	def print_results(self, n_rows_listened, n_rows_valid, n_rows_skipped, n_errors):
		print("# Extraction from Excel results:")
		print(f"{n_rows_listened} rows have been listened.")
		print(f" ├─ {n_rows_skipped} rows without data skipped.")
		print(f" └─ {n_rows_valid + n_errors} rows have been parsed.")
		if n_errors:
			print(f"     ├─ {n_rows_valid} rows have been saved.")
			print(f"     └─ {n_errors} rows are invalid! Please fix all ❌ marks first.\n")
		else:
			print("All rows with data have been saved.")
			print("No critical errors found. Going further...\n")


class SplitSalesByCountry():
	eu = dict()
	not_eu = list()

	def __init__(self, sales):
		self.all = sales
		self.eu_countries = dict()
		self.not_eu_countries = dict()
		self.split_sales()
		self.count_vat_for_eu()
		self.print_results()

	def split_sales(self):
		for row in self.all:
			country = row[1]
			if country in EU_VAT.keys():
				if country not in self.eu:
					self.eu[country] = 0
				self.eu[country] += row[2]
				if country not in self.eu_countries:
					self.eu_countries[country] = 0
				self.eu_countries[country] += 1
			else:
				self.not_eu.append(row)
				if country not in self.not_eu_countries:
					self.not_eu_countries[country] = 0
				self.not_eu_countries[country] += 1

	def count_vat_for_eu(self):
		for country in self.eu.keys():
			total = self.eu[country]
			without_vat = total * 100 / (100 + EU_VAT[country])
			vat = total * EU_VAT[country] / (100 + EU_VAT[country])
			self.eu[country] = (without_vat, vat, total)

	def print_results(self):
		print("# Sales summary:")
		self.print_countries("EU", self.eu_countries)
		self.print_countries("not EU", self.not_eu_countries)
		print()

	def print_countries(self, title, countries):
		total = 0
		for n in countries.values():
			total += n
		print(f"{total} sales in {title} countries have been found. {len(countries)} countries in total:")
		for country, n in sorted(countries.items(), key=lambda item: item[1], reverse=True):
			print(f" ● {country} - {n}")


class WriteSalesToExcel():
	def __init__(self, filename, sales, eu, not_eu):
		self.all = sales
		self.eu = eu
		self.not_eu = not_eu
		self.write(filename)

	def write(self, filename):
		workbook = Workbook()
		workbook.active.title = "Visi"
		self.write_to_sheet(workbook.active, self.all)
		self.write_to_sheet(workbook.create_sheet("ES"), [(k, *v) for k, v in self.eu.items()],
							headers=("Country", "Total without VAT", "VAT", "Total"), sum_start_from_header=2)
		self.write_to_sheet(workbook.create_sheet("ne ES"), self.not_eu)
		try:
			workbook.save(filename)
			print(f"Successfully saved sales into \"{filename}\"")
		except Exception as e:
			print(f"Failed to save sales. See the error below and close the \"{filename}\" file if it is open.")
			print(type(e), e)

	def write_to_sheet(self, sheet, sales, headers=("Date", "Country", "Total"), sum_start_from_header=3):
		sheet.append(headers)
		for cell in sheet[1]:
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal='center')
		for row in sales:
			sheet.append(row)
		self.add_sum_cells_to_sheet(sheet, len(sales), headers, sum_start_from_header)
		self.align_columns_width(sheet, (get_column_letter(len(headers)+2)))

	def add_sum_cells_to_sheet(self, sheet, n_sales, headers, start_header):
		row = 1
		for i, header in enumerate(headers[start_header-1:], start=start_header):
			row += 1
			# Header cell
			h_cell = sheet.cell(row=row, column=len(headers)+2)
			h_cell.value = header + ':'
			h_cell.font = Font(bold=True, color="FF0000")
			h_cell.alignment = Alignment(horizontal='center')
			# Sum cell
			s_cell = sheet.cell(row=row, column=len(headers)+3)
			s_cell.value = f'=SUM({get_column_letter(i)}{2}:{get_column_letter(i)}{n_sales+1})'
			s_cell.font = Font(color="C00000")

	def align_columns_width(self, sheet, columns=('E')):
		max_width = 0
		for col in columns:
			for cell in sheet[col]:
				if cell.value:
					max_width = max(max_width, len(cell.value))
			if max_width > sheet.column_dimensions[col].width:
				sheet.column_dimensions[col].width = max_width


def main():
	wb = LoadWorkbook('../EtsySoldOrders2024-7.xlsx')
	if not wb.sheet:
		return
	ext = DataExtractor(wb.sheet)
	data, status = ext.run()
	if status:
		sales = SplitSalesByCountry(data)
		EU_sales, not_EU_sales = sales.eu, sales.not_eu
		WriteSalesToExcel('sales1.xlsx', data, EU_sales, not_EU_sales)

if __name__ == '__main__':
	main()
